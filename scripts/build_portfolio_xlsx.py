#!/usr/bin/env python3
"""Build portfolio.xlsx that EXACTLY mirrors Shawn's Edward Jones LIRA statements
(account 503-18809), Canadian + U.S. sleeves, as of May 31 2026. Every line ties
to the PDF so the total is verifiable. Cost basis = Adjusted Cost Base (ACB) from
the statement. USD->CAD at 1.41."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

USDCAD = 1.41

# name, ticker, shares, total_acb, ej_price(May31), ej_market_value  -- all from statement
CAD_STOCKS = [
    ("Bank of Nova Scotia", "BNS", 180,      12710.05, 110.62, 19911.60),
    ("Enbridge",            "ENB", 275,      13954.28,  75.64, 20801.00),
    ("Manulife Financial",  "MFC", 500,       8250.59,  52.73, 26365.00),
    ("MDA Space",           "MDA", 900,      40079.52,  61.48, 55332.00),
    ("TC Energy",           "TRP", 250,      12180.32,  91.86, 22965.00),
    ("Toronto-Dominion",    "TD",  200,       9838.32, 157.75, 31550.00),
]
CAD_FUNDS = [
    ("AGF Elements Balanced LL",        "AGF498",   1966.656,  22908.20, 16.186, 31832.29),
    ("BMO Asian Growth & Income DSC",   "BMO22120", 1454.2107, 23074.75, 32.991, 47975.87),
    ("BMO Monthly High Income II",      "BMO22619",  928.832,   8273.94, 17.629, 16374.38),
    ("Franklin Royce Gbl SmCap Prem A", "TML707",   1414.1788, 25896.02, 29.752, 42074.65),
    ("Mackenzie Cdn Equity Sr A",       "MFC2946",  1229.483,  25523.45, 49.559, 60931.95),
]
# Money market + cash (Canadian sleeve)
CAD_MM   = ("EVF High Interest Savings A", "EVF", 8670, 86700.00, 10.00, 86700.00)
CAD_CASH = 140.46

# U.S. sleeve (USD) -- name, ticker, shares, total_acb, ej_price, ej_market_value
US_STOCKS = [
    ("ASML Holding",     "ASML", 5,   6908.26, 1612.76,  8063.80),
    ("BWX Technologies", "BWXT", 125, 24910.14,  195.88, 24485.00),
    ("ServiceNow",       "NOW",  265, 25136.85,  124.37, 32958.05),
]
US_CASH = 13.47

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "LIRA"
NAVY="1F3A5F"; LBLUE="D6E4F0"; GREEN="0A7A4F"; RED="C0392B"
wfont=Font(color="FFFFFF", bold=True)
hdr_fill=PatternFill("solid", fgColor=NAVY); sec_fill=PatternFill("solid", fgColor=LBLUE)
thin=Side(style="thin", color="CCCCCC"); border=Border(bottom=thin)
moneyc='$#,##0'; moneyc2='$#,##0.00'

ws.merge_cells("A1:H1")
c=ws["A1"]; c.value="MY EDWARD JONES LIRA  (acct 503-18809) — ties to your statements"
c.font=Font(bold=True, size=15, color="FFFFFF"); c.fill=hdr_fill
c.alignment=Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height=26
ws.merge_cells("A2:H2")
ws["A2"]="Holdings as of May 31, 2026 (statement prices) · cost = ACB from statement · USD→CAD 1.41"
ws["A2"].font=Font(italic=True, size=9, color="666666")

headers=["Holding","Ticker","Shares","Cost (ACB)","Price","Market Value","Value CAD","Gain/Loss CAD"]
r=4
for i,h in enumerate(headers,1):
    cell=ws.cell(r,i,h); cell.font=wfont; cell.fill=hdr_fill
    cell.alignment=Alignment(horizontal="right" if i>2 else "left")

grand_cad=0.0; grand_cost_cad=0.0

def section(title):
    global r
    r+=1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    sc=ws.cell(r,1,title); sc.font=Font(bold=True, color=NAVY); sc.fill=sec_fill

def row(name, tk, shares, acb, price, mv, fx):
    global r, grand_cad, grand_cost_cad
    r+=1
    val_cad=mv*fx; cost_cad=acb*fx; gain_cad=val_cad-cost_cad
    grand_cad+=val_cad; grand_cost_cad+=cost_cad
    ws.cell(r,1,name); ws.cell(r,2,tk); ws.cell(r,3,shares)
    ws.cell(r,4,acb).number_format=moneyc
    ws.cell(r,5,price).number_format=(moneyc2 if price<100 else moneyc)
    ws.cell(r,6,mv).number_format=moneyc
    ws.cell(r,7,val_cad).number_format=moneyc
    g=ws.cell(r,8,gain_cad); g.number_format=moneyc
    g.font=Font(color=GREEN if gain_cad>=0 else RED, bold=True)
    for cc in range(1,9): ws.cell(r,cc).border=border

section("🇨🇦 CANADIAN STOCKS")
for n,tk,sh,acb,p,mv in CAD_STOCKS: row(n,tk,sh,acb,p,mv,1.0)
section("🇨🇦 MONEY MARKET (your safe cash sleeve)")
n,tk,sh,acb,p,mv=CAD_MM; row(n,tk,sh,acb,p,mv,1.0)
section("🇨🇦 MUTUAL FUNDS (the 5 you're reviewing)")
for n,tk,sh,acb,p,mv in CAD_FUNDS: row(n,tk,sh,acb,p,mv,1.0)
section("🇺🇸 U.S. STOCKS (shown in USD, converted at 1.41)")
for n,tk,sh,acb,p,mv in US_STOCKS: row(n,tk,sh,acb,p,mv,USDCAD)

# Cash lines
r+=1; ws.cell(r,1,"Cash (CAD sleeve)"); ws.cell(r,7,CAD_CASH).number_format=moneyc
grand_cad+=CAD_CASH; grand_cost_cad+=CAD_CASH
for cc in range(1,9): ws.cell(r,cc).border=border
r+=1; ws.cell(r,1,"Cash (US sleeve)"); usc=US_CASH*USDCAD
ws.cell(r,7,usc).number_format=moneyc; grand_cad+=usc; grand_cost_cad+=usc
for cc in range(1,9): ws.cell(r,cc).border=border

# Total
r+=2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
tc=ws.cell(r,1,"TOTAL LIRA (CAD)"); tc.font=Font(bold=True, size=13, color="FFFFFF"); tc.fill=hdr_fill
tv=ws.cell(r,7,grand_cad); tv.number_format=moneyc; tv.font=Font(bold=True, size=13, color="FFFFFF"); tv.fill=hdr_fill
ws.cell(r,8,"").fill=hdr_fill; ws.row_dimensions[r].height=24
r+=1
gc=ws.cell(r,1,"Total unrealized gain"); gc.font=Font(bold=True)
gtot=grand_cad-grand_cost_cad
gv=ws.cell(r,8,gtot); gv.number_format=moneyc; gv.font=Font(bold=True, color=GREEN if gtot>=0 else RED)

# Verify ties
r+=2
ties=[
    "✅ Ties to your statements: CAD sleeve $462,954.20 + US sleeve US$65,520.32. Every line above matches your EJ PDF.",
    "⚠️ NOT in these statements: Rocket Lab (RKLB) and Beam (BEAM). If you own them, they're in your RRSP or Canadian Investment account — send me that statement and I'll add them.",
    "ℹ️ Prices are May 31 statement prices. Since then MDA is ~$57.80 (down from $61.48), ASML up. I anchor to the statement so you can verify; I'll layer live prices once you trust this baseline.",
]
for n in ties:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r,1,n).font=Font(size=9, color="555555"); r+=1

widths=[28,9,10,12,10,14,13,14]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="A5"
wb.save("/home/user/Claude-md/portfolio.xlsx")
print(f"saved. TOTAL LIRA CAD = {grand_cad:,.2f} | cost = {grand_cost_cad:,.2f} | gain = {gtot:,.2f}")
